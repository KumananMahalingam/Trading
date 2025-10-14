import time
import secret_key
from openpyxl import Workbook
from datetime import datetime, timezone
import spacy
from polygon import RESTClient
from collections import defaultdict, Counter
import json
import re
import pandas as pd
import yfinance as yf
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from groq import Groq

client = RESTClient(secret_key.API_KEY)
groq_client = Groq(api_key=secret_key.GROQ_API_KEY)
nlp = spacy.load("en_core_web_sm")
sia = SentimentIntensityAnalyzer()

companies = {
    "AAPL": "Apple",
    "JPM": "JPMorgan Chase & Co",
    "PEP": "Pepsi",
    "TM": "Toyota",
    "AMZN": "Amazon"
}

# Company name similarity mapping to avoid duplicates
COMPANY_ALIASES = {
    "APPLE": ["AAPL"],
    "JPMORGAN": ["JPM", "JPM.N"],
    "PEPSI": ["PEP", "PEPSICO"],
    "TOYOTA": ["TM", "TYO"],
    "AMAZON": ["AMZN"]
}

def load_company_tickers_json(file_path="company_tickers.json"):
    """
    Load the SEC company tickers JSON and create lookup dictionaries
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        print(f"Loaded {len(data)} companies from {file_path}")

        ticker_to_info = {}
        name_to_ticker = {}
        name_variations = {}

        for key, company in data.items():
            ticker = company.get('ticker', '').upper()
            name = company.get('title', '')
            cik = company.get('cik_str', '')

            if ticker and name:
                ticker_to_info[ticker] = {
                    'name': name,
                    'ticker': ticker,
                    'cik': cik
                }

                name_to_ticker[name.upper()] = ticker

                variations = generate_name_variations(name)
                for variation in variations:
                    if len(variation) >= 3:
                        name_variations[variation.upper()] = ticker

        print(f"Created {len(name_variations)} name variations for matching")

        return {
            'ticker_to_info': ticker_to_info,
            'name_to_ticker': name_to_ticker,
            'name_variations': name_variations
        }

    except FileNotFoundError:
        print(f"Error: {file_path} not found. Please download it first.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON: {e}")
        return None

def generate_name_variations(company_name):
    """
    Generate common variations of company names for better matching
    """
    variations = [company_name]

    suffixes_pattern = r'\s+(Inc\.?|Corporation|Corp\.?|Company|Co\.?|Limited|Ltd\.?|Holdings|Holding|LLC|L\.L\.C\.|plc|PLC)'
    base_name = re.sub(suffixes_pattern, '', company_name, flags=re.IGNORECASE).strip()
    variations.append(base_name)

    variations.extend([
        company_name.replace(',', '').strip(),
        company_name.replace('.', '').strip(),
        base_name.replace(',', '').strip(),
        base_name.replace('.', '').strip()
    ])

    variations.extend([
        company_name.replace('&', 'and'),
        company_name.replace(' and ', ' & '),
        base_name.replace('&', 'and'),
        base_name.replace(' and ', ' & ')
    ])

    words = base_name.split()
    if len(words) > 1:
        acronym = ''.join([word[0] for word in words if word[0].isupper()])
        if len(acronym) >= 2:
            variations.append(acronym)

    return list(set([v.strip() for v in variations if v.strip()]))

def validate_company_exists(company_name, company_lookups):
    """
    Check if a company name corresponds to a real publicly traded company
    Returns tuple: (is_valid, ticker, official_name)
    """
    if not company_lookups:
        return False, None, None

    company_upper = company_name.upper().strip()

    if company_upper in company_lookups['name_to_ticker']:
        ticker = company_lookups['name_to_ticker'][company_upper]
        official_name = company_lookups['ticker_to_info'][ticker]['name']
        return True, ticker, official_name

    if company_upper in company_lookups['name_variations']:
        ticker = company_lookups['name_variations'][company_upper]
        official_name = company_lookups['ticker_to_info'][ticker]['name']
        return True, ticker, official_name

    if company_upper in company_lookups['ticker_to_info']:
        ticker = company_upper
        official_name = company_lookups['ticker_to_info'][ticker]['name']
        return True, ticker, official_name

    for variation, ticker in company_lookups['name_variations'].items():
        if len(company_upper) >= 4 and company_upper in variation:
            official_name = company_lookups['ticker_to_info'][ticker]['name']
            return True, ticker, official_name
        elif len(variation) >= 4 and variation in company_upper:
            official_name = company_lookups['ticker_to_info'][ticker]['name']
            return True, ticker, official_name

    return False, None, None

def is_same_company(ticker1, ticker2, name1, name2):
    """
    Check if two tickers represent the same company
    """
    if ticker1 == ticker2:
        return True

    # Check company aliases
    for base_name, aliases in COMPANY_ALIASES.items():
        if ticker1 in aliases and ticker2 in aliases:
            return True

    # Check name similarity (basic check)
    name1_base = re.sub(r'\s+(Inc\.?|Corporation|Corp\.?|Company|Co\.?|Limited|Ltd\.?)', '', name1, flags=re.IGNORECASE).strip().upper()
    name2_base = re.sub(r'\s+(Inc\.?|Corporation|Corp\.?|Company|Co\.?|Limited|Ltd\.?)', '', name2, flags=re.IGNORECASE).strip().upper()

    # If base names are very similar
    if name1_base == name2_base:
        return True

    return False

def validate_ticker_quality(ticker, min_market_cap=1e9, required_exchanges=['NYSE', 'NASDAQ', 'NYQ', 'NMS']):
    """
    Validate that a ticker meets quality criteria:
    - Listed on major US exchange
    - Has sufficient market cap
    - Has valid trading data

    Returns: (is_valid, reason)
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        # Check if data exists
        if not info or len(info) < 5:
            return False, "Insufficient data"

        # Check exchange
        exchange = info.get('exchange', '').upper()
        if exchange not in required_exchanges:
            return False, f"Not on major exchange (found: {exchange})"

        # Check market cap
        market_cap = info.get('marketCap', 0)
        if market_cap < min_market_cap:
            return False, f"Market cap too low (${market_cap:,.0f})"

        # Check if actively traded
        volume = info.get('volume', 0)
        if volume < 100000:  # Less than 100k daily volume
            return False, "Low trading volume"

        return True, "Valid"

    except Exception as e:
        return False, f"Error: {str(e)}"

def fetch_stock_data(ticker, start_date, end_date):
    """
    Fetch historical stock data from Yahoo Finance
    """
    try:
        print(f"  Fetching stock data for {ticker}...")

        # Convert dates to proper format
        start = start_date.split('T')[0]
        end = end_date.split('T')[0]

        # Download data from yfinance
        stock = yf.Ticker(ticker)
        df = stock.history(start=start, end=end)

        if df.empty:
            print(f"  No data available for {ticker}")
            return pd.DataFrame()

        # Rename columns to lowercase and reset index
        df.reset_index(inplace=True)
        df.columns = df.columns.str.lower()

        # Keep only necessary columns and rename Date to date
        df = df.rename(columns={'date': 'date'})
        df['date'] = df['date'].dt.strftime('%Y-%m-%d')

        # Select relevant columns (yfinance returns: Date, Open, High, Low, Close, Volume)
        df = df[['date', 'open', 'high', 'low', 'close', 'volume']]

        print(f"  Fetched {len(df)} days of stock data for {ticker}")
        return df

    except Exception as e:
        print(f"  Error fetching stock data for {ticker}: {e}")
        return pd.DataFrame()

def calculate_technical_indicators(df):
    """
    Calculate technical indicators for stock data
    """
    if df.empty:
        return df

    # Simple Moving Averages
    df['SMA_5'] = df['close'].rolling(window=5).mean()
    df['SMA_20'] = df['close'].rolling(window=20).mean()
    df['SMA_50'] = df['close'].rolling(window=50).mean()

    # Exponential Moving Averages
    df['EMA_12'] = df['close'].ewm(span=12, adjust=False).mean()
    df['EMA_26'] = df['close'].ewm(span=26, adjust=False).mean()

    # MACD
    df['MACD'] = df['EMA_12'] - df['EMA_26']
    df['MACD_Signal'] = df['MACD'].ewm(span=9, adjust=False).mean()

    # RSI
    delta = df['close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    df['RSI'] = 100 - (100 / (1 + rs))

    # Bollinger Bands
    df['BB_Middle'] = df['close'].rolling(window=20).mean()
    bb_std = df['close'].rolling(window=20).std()
    df['BB_Upper'] = df['BB_Middle'] + (bb_std * 2)
    df['BB_Lower'] = df['BB_Middle'] - (bb_std * 2)

    # Returns
    df['Return_1D'] = df['close'].pct_change(1)
    df['Return_5D'] = df['close'].pct_change(5)
    df['Return_20D'] = df['close'].pct_change(20)

    # Volatility
    df['Volatility_20D'] = df['Return_1D'].rolling(window=20).std()

    return df

def prepare_dataframe_for_alpha(ticker, stock_df, daily_sentiments, related_companies):
    """
    Prepare comprehensive DataFrame with stock data, technical indicators, and sentiment
    """
    if stock_df.empty:
        return None

    # Calculate technical indicators
    df = calculate_technical_indicators(stock_df.copy())

    # Add sentiment data for target company
    sentiment_dict = {}
    for date, scores in daily_sentiments[ticker].items():
        sentiment_dict[date] = sum(scores) / len(scores) if scores else 0

    df[f'{ticker}_Sentiment'] = df['date'].map(sentiment_dict).fillna(0)

    # Add sentiment data for related companies
    for related_ticker in related_companies:
        if related_ticker in daily_sentiments:
            related_sentiment_dict = {}
            for date, scores in daily_sentiments[related_ticker].items():
                related_sentiment_dict[date] = sum(scores) / len(scores) if scores else 0

            df[f'{related_ticker}_Sentiment'] = df['date'].map(related_sentiment_dict).fillna(0)

    # Add sentiment divergence features
    for related_ticker in related_companies:
        if f'{related_ticker}_Sentiment' in df.columns:
            df[f'Sentiment_Div_{ticker}_{related_ticker}'] = df[f'{ticker}_Sentiment'] - df[f'{related_ticker}_Sentiment']

    return df

def generate_predictive_alphas(ticker, company_name, dataframe_json, related_companies):
    """
    Use Deepseek R1 to generate predictive alpha formulas
    """
    try:
        # Get list of related company tickers for the prompt
        related_list = ", ".join(related_companies) if related_companies else "None available"

        prompt = f"""Generating Predictive Alphas for {company_name} ({ticker}) Stock Prices

        Objective: Generate formulaic alpha signals to predict {company_name}'s stock prices using:
        1. Stock features for {ticker}
        2. Technical indicators for {ticker}
        3. Sentiment data for {ticker} and related companies

        Related Companies Available: {related_list}

        Input Data: A single pandas.DataFrame with rows representing trading days and columns including:
        - Stock Features for {ticker}: close, open, high, low, volume
        - Technical Indicators for {ticker}: RSI, SMA_5, SMA_20, SMA_50, EMA_12, EMA_26, MACD, MACD_Signal, BB_Upper, BB_Middle, BB_Lower
        - Sentiment Data: {ticker}_Sentiment, {', '.join([f'{rt}_Sentiment' for rt in related_companies]) if related_companies else 'None'}
        - Returns for {ticker}: Return_1D, Return_5D, Return_20D
        - Volatility for {ticker}: Volatility_20D

        DataFrame Sample (last 10 rows):
        {dataframe_json}

        Requirements:
        1. Generate 5 SPECIFIC alpha formulas using ONLY the column names that exist in the DataFrame above
        2. Use actual ticker symbols (e.g., {ticker}_Sentiment, GM_Sentiment, TSLA_Sentiment) NOT placeholders like "RelatedCompany"
        3. Each formula must use only features present in the DataFrame columns
        4. Focus on:
           - Momentum: Return_5D, Return_20D
           - Mean reversion: RSI, Bollinger Bands
           - Sentiment divergences: {ticker}_Sentiment minus specific related company sentiments
           - Technical crossovers: MACD, moving average crosses

        Example format (use ACTUAL ticker symbols):
        α1 = TM_Return_5D + 0.5 × (TM_Sentiment - GM_Sentiment)

        Provide ONLY 5 implementable formulas with brief explanations."""

        response = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "You are a quantitative analyst. Generate SPECIFIC alpha formulas using ONLY the exact column names provided. Never use generic placeholders like 'RelatedCompany'."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5,  # Lower temperature for more consistent output
            max_tokens=2000
        )

        alphas = response.choices[0].message.content
        return alphas

    except Exception as e:
        print(f"Error generating alphas for {ticker}: {e}")
        return None

print("Loading company tickers JSON...")
company_lookups = load_company_tickers_json("company_tickers.json")

if not company_lookups:
    print("Warning: Could not load company data. Continuing with basic analysis...")

wb = Workbook()
sheet = wb.active
sheet.title = "News Data"
sheet.append(["Company", "Ticker", "Date", "Headline", "URL", "Summary", "Sentiment Score"])

daily_sheet = wb.create_sheet("Daily Sentiment")
daily_sheet.append(["Company", "Ticker", "Date", "Average Sentiment", "Article Count"])

alpha_sheet = wb.create_sheet("Predictive Alphas")
alpha_sheet.append(["Company", "Ticker", "Related Companies", "Generated Alphas"])

start_date = datetime(2023, 9, 1, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
end_date   = datetime(2025, 9, 1, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

def fetch_news(ticker, start, end, batch_size=1000, sleep_time=12):
    """
    Fetch news for a ticker with pagination & backoff handling
    """
    results = []
    batch_count = 0

    try:
        response = client.list_ticker_news(
            ticker=ticker,
            published_utc_gte=start_date,
            published_utc_lte=end_date,
            limit=batch_size,
            order="asc"
        )

        current_batch = []
        for item in response:
            current_batch.append(item)
            if len(current_batch) >= batch_size:
                break

        while current_batch:
            results.extend(current_batch)
            batch_count += 1
            print(f"  Fetched batch {batch_count} ({len(current_batch)} articles) for {ticker}. Total: {len(results)}")

            if len(current_batch) < batch_size:
                print(f"  Reached end of articles for {ticker}")
                break

            last_article_date = current_batch[-1].published_utc

            if last_article_date >= end_date:
                print(f"  Reached end date for {ticker}")
                break

            print(f"  Waiting {sleep_time} seconds before next batch...")
            time.sleep(sleep_time)

            response = client.list_ticker_news(
                ticker=ticker,
                published_utc_gt=last_article_date,
                published_utc_lte=end_date,
                limit=batch_size,
                order='asc'
            )

            current_batch = []
            for item in response:
                current_batch.append(item)
                if len(current_batch) >= batch_size:
                    break

        print(f"  Total articles fetched for {ticker}: {len(results)}")

    except Exception as e:
        print(f"Error fetching news for {ticker}: {e}")

    return results

news_articles = []
daily_sentiments = defaultdict(lambda: defaultdict(list))
mentions_by_source = defaultdict(Counter)
validated_mentions = defaultdict(Counter)
invalid_mentions = defaultdict(Counter)
stock_dataframes = {}

# First pass: Fetch news and sentiment data
for ticker, name in companies.items():
    print(f"Fetching news for {name} ({ticker})...")
    news_articles = fetch_news(ticker, start_date, end_date, batch_size=1000, sleep_time=12)

    for item in news_articles:
        text = f"{item.title} {item.summary if hasattr(item, 'summary') else ''}"
        sentiment_scores = sia.polarity_scores(text)
        compound_score = sentiment_scores['compound']

        article_date = item.published_utc.split('T')[0] if 'T' in item.published_utc else item.published_utc[:10]
        daily_sentiments[ticker][article_date].append(compound_score)

        doc = nlp(text)
        for ent in doc.ents:
            if ent.label_ == "ORG":
                mentions_by_source[ticker][ent.text] += 1

                if company_lookups:
                    is_valid, mention_ticker, official_name = validate_company_exists(ent.text, company_lookups)
                    if is_valid:
                        validated_mentions[ticker][f"{official_name} ({mention_ticker})"] += 1
                    else:
                        invalid_mentions[ticker][ent.text] += 1

        row = [
            name,
            ticker,
            item.published_utc,
            item.title,
            item.article_url,
            item.summary if hasattr(item, "summary") else "",
            compound_score
        ]
        sheet.append(row)

    print(f"Waiting 65 seconds before next company...")
    time.sleep(65)

# Populate Daily Sentiment sheet
print("\nCreating daily sentiment aggregations...")
for ticker in sorted(daily_sentiments.keys()):
    company_name = companies.get(ticker, ticker)
    for date in sorted(daily_sentiments[ticker].keys()):
        scores = daily_sentiments[ticker][date]
        avg_sentiment = sum(scores) / len(scores)
        article_count = len(scores)

        daily_sheet.append([
            company_name,
            ticker,
            date,
            avg_sentiment,
            article_count
        ])

# Second pass: Fetch stock data and generate alphas
print("\n" + "="*80)
print("FETCHING STOCK DATA AND GENERATING PREDICTIVE ALPHAS")
print("="*80)

for ticker, name in companies.items():
    print(f"\nProcessing {name} ({ticker})...")

    # Fetch stock data
    stock_df = fetch_stock_data(ticker, start_date, end_date)

    if not stock_df.empty:
        stock_dataframes[ticker] = stock_df

        # Identify related companies with enhanced filtering
        related_companies = []
        candidates_checked = 0
        max_candidates = 20  # Check more candidates to account for filtering

        print(f"  Identifying related companies for {ticker}...")

        for company_str, count in validated_mentions[ticker].most_common(max_candidates):
            if len(related_companies) >= 5:  # Stop after finding 5 valid ones
                break

            candidates_checked += 1

            # Extract ticker from "Company Name (TICKER)" format
            match = re.search(r'\(([A-Z]+)\)', company_str)
            if not match:
                continue

            related_ticker = match.group(1)

            # Skip if same ticker
            if related_ticker == ticker:
                print(f"    ✗ Skipped {related_ticker}: Same as target company")
                continue

            # Extract company name
            related_name = company_str.split('(')[0].strip()

            # Check if it's the same company with different ticker
            if is_same_company(ticker, related_ticker, name, related_name):
                print(f"    ✗ Skipped {related_ticker}: Same company, different ticker")
                continue

            # Validate ticker quality
            is_valid, reason = validate_ticker_quality(related_ticker, min_market_cap=5e9)
            if not is_valid:
                print(f"    ✗ Skipped {related_ticker}: {reason}")
                continue

            # All checks passed
            related_companies.append(related_ticker)
            print(f"    ✓ Added {related_ticker} ({related_name}): {count} mentions")

        print(f"  Found {len(related_companies)} valid related companies (checked {candidates_checked} candidates)")

        # Prepare comprehensive DataFrame
        comprehensive_df = prepare_dataframe_for_alpha(
            ticker,
            stock_df,
            daily_sentiments,
            related_companies
        )

        if comprehensive_df is not None and not comprehensive_df.empty:
            # Convert last 10 rows to JSON for the prompt
            sample_df = comprehensive_df.tail(10).fillna(0)
            df_json = sample_df.to_json(orient='records', indent=2)

            # Generate predictive alphas
            print(f"  Generating predictive alphas for {ticker}...")
            alphas = generate_predictive_alphas(
                ticker,
                name,
                df_json,
                related_companies
            )

            if alphas:
                alpha_sheet.append([
                    name,
                    ticker,
                    ", ".join(related_companies) if related_companies else "None",
                    alphas
                ])

                print(f"\n{'='*80}")
                print(f"PREDICTIVE ALPHAS FOR {name} ({ticker})")
                print(f"Related Companies: {', '.join(related_companies) if related_companies else 'None'}")
                print(f"{'='*80}")
                print(alphas)
                print(f"{'='*80}\n")

        time.sleep(2)

wb.save("news.xlsx")
print("\n" + "="*80)
print("Analysis complete! Saved to news.xlsx with predictive alphas")
print("="*80)

# Print summary statistics
if company_lookups:
    print("\n" + "="*80)
    print("VALIDATED PUBLIC COMPANIES MENTIONED")
    print("="*80)

    for ticker, counter in validated_mentions.items():
        company_name = companies.get(ticker, ticker)
        print(f"\n{company_name} ({ticker}) - Top mentions:")
        for company, count in counter.most_common(10):
            print(f"  ✓ {company}: {count} mentions")