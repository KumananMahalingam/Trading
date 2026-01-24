"""
Data caching management
"""
import os
from openpyxl import load_workbook


def check_essential_data_only(file_path, tickers):
    """
    Check only for daily sentiments, stock data, and related companies

    Args:
        file_path: Path to Excel file
        tickers: List of ticker symbols to check

    Returns:
        bool: True if data is complete, False otherwise
    """
    try:
        if not os.path.exists(file_path):
            return False

        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames

        if 'Daily Sentiment' not in sheet_names:
            print("  Missing: Daily Sentiment sheet")
            wb.close()
            return False

        if 'Related Companies' not in sheet_names:
            print("  Missing: Related Companies sheet")
            wb.close()
            return False

        if 'Stock Prices' not in sheet_names:
            print("  Missing: Stock Prices sheet")
            wb.close()
            return False

        sentiments_sheet = wb['Daily Sentiment']
        related_sheet = wb['Related Companies']
        stock_sheet = wb['Stock Prices']

        sentiment_rows = list(sentiments_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        related_rows = list(related_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        stock_rows = list(stock_sheet.iter_rows(min_row=2, values_only=True))

        if not sentiment_rows or not sentiment_rows[0][0]:
            print("  Daily Sentiment sheet is empty")
            wb.close()
            return False

        if not related_rows or not related_rows[0][0]:
            print("  Related Companies sheet is empty")
            wb.close()
            return False

        if not stock_rows or not stock_rows[0][0]:
            print("  Stock Prices sheet is empty")
            wb.close()
            return False

        tickers_in_file = set()
        for row in stock_rows:
            if len(row) >= 1 and row[0]:
                tickers_in_file.add(row[0])

        wb.close()

        missing_tickers = set(tickers) - tickers_in_file
        if missing_tickers:
            print(f"  Missing stock data for: {', '.join(missing_tickers)}")
            return False

        print("  ✓ All essential data found in cache")
        return True

    except Exception as e:
        print(f"  Error checking cache: {e}")
        import traceback
        traceback.print_exc()
        return False


def check_data_completeness(filename, required_tickers):
    """
    Check if Excel file has complete data for all required tickers

    Args:
        filename: Excel file path
        required_tickers: list of ticker symbols

    Returns:
        bool: True if data is complete, False otherwise
    """
    if not os.path.exists(filename):
        return False

    try:
        wb = load_workbook(filename, read_only=True)

        if "Metadata" not in wb.sheetnames:
            wb.close()
            return False

        meta_sheet = wb["Metadata"]
        meta_rows = list(meta_sheet.iter_rows(min_row=2, values_only=True))

        for row in meta_rows:
            if len(row) >= 2 and row[0] == "Data Complete":
                data_complete = row[1] == "Yes"
                wb.close()

                if not data_complete:
                    return False

        if "Stock Prices" not in wb.sheetnames:
            wb.close()
            return False

        stock_sheet = wb["Stock Prices"]
        stock_rows = list(stock_sheet.iter_rows(min_row=2, values_only=True))

        tickers_in_file = set()
        for row in stock_rows:
            if len(row) >= 1 and row[0]:
                tickers_in_file.add(row[0])

        wb.close()

        missing_tickers = set(required_tickers) - tickers_in_file

        if missing_tickers:
            print(f"Missing data for tickers: {', '.join(missing_tickers)}")
            return False

        return True

    except Exception as e:
        print(f"Error checking data completeness: {e}")
        return False


def display_data_summary(daily_sentiments, stock_dataframes, all_related_companies):
    """
    Display a summary of loaded data
    """
    print(f"\n{'='*80}")
    print("DATA SUMMARY")
    print(f"{'='*80}")

    print(f"\nSentiment Data:")
    for ticker in sorted(daily_sentiments.keys())[:10]:
        days = len(daily_sentiments[ticker])
        print(f"  {ticker}: {days} days")
    if len(daily_sentiments) > 10:
        print(f"  ... and {len(daily_sentiments) - 10} more tickers")

    print(f"\nStock Price Data:")
    for ticker in sorted(stock_dataframes.keys()):
        rows = len(stock_dataframes[ticker])
        print(f"  {ticker}: {rows} days")

    print(f"\nRelated Companies:")
    for ticker in sorted(all_related_companies.keys()):
        related = all_related_companies[ticker]
        print(f"  {ticker}: {', '.join(related)}")

    print(f"\n{'='*80}\n")