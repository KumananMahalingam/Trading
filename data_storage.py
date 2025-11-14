import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from collections import defaultdict, Counter
import json
import re


def save_all_data_to_excel(
    filename,
    daily_sentiments,
    stock_dataframes,
    validated_mentions,
    all_related_companies,
    alpha_texts,
    trained_models=None
):
    """
    Save all pipeline data to a comprehensive Excel file

    Args:
        filename: Excel file path
        daily_sentiments: dict of {ticker: {date: [scores]}}
        stock_dataframes: dict of {ticker: DataFrame}
        validated_mentions: dict of {ticker: Counter}
        all_related_companies: dict of {ticker: [related_tickers]}
        alpha_texts: dict of {ticker: alpha_text}
        trained_models: dict of {ticker: {'metrics': {...}}}
    """

    print(f"\n{'='*80}")
    print(f"SAVING ALL DATA TO {filename}")
    print(f"{'='*80}")

    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    print("  Saving daily sentiment data...")
    sentiment_sheet = wb.create_sheet("Daily Sentiment")
    sentiment_sheet.append(["Ticker", "Date", "Average Sentiment", "Article Count", "Raw Scores"])

    sentiment_count = 0
    for ticker in sorted(daily_sentiments.keys()):
        for date in sorted(daily_sentiments[ticker].keys()):
            scores = daily_sentiments[ticker][date]
            avg_sentiment = sum(scores) / len(scores) if scores else 0
            article_count = len(scores)
            # Store raw scores as JSON string for exact reconstruction
            raw_scores_json = json.dumps(scores)
            sentiment_sheet.append([ticker, date, avg_sentiment, article_count, raw_scores_json])
            sentiment_count += 1

    print(f"    Saved {sentiment_count} sentiment records")
    print("  Saving stock price data...")
    stock_sheet = wb.create_sheet("Stock Prices")
    stock_sheet.append(["Ticker", "Date", "Open", "High", "Low", "Close", "Volume"])

    stock_count = 0
    for ticker, df in stock_dataframes.items():
        for _, row in df.iterrows():
            stock_sheet.append([
                ticker,
                row['date'],
                row['open'],
                row['high'],
                row['low'],
                row['close'],
                row['volume']
            ])
            stock_count += 1

    print(f"    Saved {stock_count} stock price records")

    print("  Saving related companies...")
    related_sheet = wb.create_sheet("Related Companies")
    related_sheet.append(["Primary Ticker", "Related Ticker", "Mention Count", "Rank"])

    for ticker, mentions in validated_mentions.items():
        for rank, (company_str, count) in enumerate(mentions.most_common(20), 1):
            # Extract ticker from "Company Name (TICKER)" format
            match = re.search(r'\(([A-Z]+)\)', company_str)
            if match:
                related_ticker = match.group(1)
                related_sheet.append([ticker, related_ticker, count, rank])

    print(f"    Saved related companies for {len(validated_mentions)} tickers")
    print("  Saving alpha formulas...")
    alpha_sheet = wb.create_sheet("Alpha Formulas")
    alpha_sheet.append(["Ticker", "Related Companies", "Alpha Text"])

    for ticker, alpha_text in alpha_texts.items():
        related = ", ".join(all_related_companies.get(ticker, []))
        alpha_sheet.append([ticker, related, alpha_text])

    print(f"    Saved alpha formulas for {len(alpha_texts)} tickers")
    print("  Saving model results...")
    results_sheet = wb.create_sheet("Model Results")
    results_sheet.append([
        "Ticker", "MSE", "RMSE", "MAE", "MAPE",
        "Directional Accuracy", "Within 1% Accuracy"
    ])

    if trained_models:
        for ticker, model_data in trained_models.items():
            metrics = model_data['metrics']
            results_sheet.append([
                ticker,
                metrics.get('MSE', 0),
                metrics.get('RMSE', 0),
                metrics.get('MAE', 0),
                metrics.get('MAPE', 0) if metrics.get('MAPE') != float('inf') else 'N/A',
                metrics.get('Directional Accuracy', 0),
                metrics.get('Within 1% Accuracy', 0)
            ])

    print(f"    Saved model results for {len(trained_models) if trained_models else 0} models")

    print("  Saving metadata...")
    meta_sheet = wb.create_sheet("Metadata")
    meta_sheet.append(["Key", "Value"])
    meta_sheet.append(["Total Tickers", len(daily_sentiments)])
    meta_sheet.append(["Primary Companies", len(stock_dataframes)])
    meta_sheet.append(["Total Related Companies", len(all_related_companies)])
    meta_sheet.append(["Data Complete", "Yes"])

    from datetime import datetime
    meta_sheet.append(["Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    wb.save(filename)
    print(f"\n✓ All data saved to {filename}")
    print(f"{'='*80}\n")


def load_all_data_from_excel(filename):
    """
    Load all pipeline data from Excel file

    Returns:
        tuple: (daily_sentiments, stock_dataframes, validated_mentions,
                all_related_companies, alpha_texts, success)
    """

    if not os.path.exists(filename):
        print(f"❌ File {filename} not found")
        return None, None, None, None, None, False

    print(f"\n{'='*80}")
    print(f"LOADING ALL DATA FROM {filename}")
    print(f"{'='*80}")

    try:
        wb = load_workbook(filename, read_only=True)

        required_sheets = ["Daily Sentiment", "Stock Prices", "Related Companies",
                          "Alpha Formulas", "Metadata"]

        missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
        if missing_sheets:
            print(f"❌ Missing required sheets: {', '.join(missing_sheets)}")
            wb.close()
            return None, None, None, None, None, False

        print("  Loading daily sentiment data...")
        daily_sentiments = defaultdict(lambda: defaultdict(list))

        sentiment_sheet = wb["Daily Sentiment"]
        sentiment_rows = list(sentiment_sheet.iter_rows(min_row=2, values_only=True))

        for row in sentiment_rows:
            if len(row) >= 5:
                ticker, date, avg_sentiment, article_count, raw_scores_json = row[:5]
                if ticker and date:
                    # Reconstruct exact scores from JSON
                    try:
                        scores = json.loads(raw_scores_json)
                        daily_sentiments[ticker][date] = scores
                    except:
                        scores = [avg_sentiment] * int(article_count)
                        daily_sentiments[ticker][date] = scores

        print(f"    Loaded sentiment for {len(daily_sentiments)} tickers")

        print("  Loading stock price data...")
        stock_dataframes = {}

        stock_sheet = wb["Stock Prices"]
        stock_rows = list(stock_sheet.iter_rows(min_row=2, values_only=True))

        # Group by ticker
        stock_data_by_ticker = defaultdict(list)
        for row in stock_rows:
            if len(row) >= 7:
                ticker, date, open_p, high, low, close, volume = row[:7]
                if ticker and date:
                    stock_data_by_ticker[ticker].append({
                        'date': date,
                        'open': open_p,
                        'high': high,
                        'low': low,
                        'close': close,
                        'volume': volume
                    })

        # Convert to DataFrames
        for ticker, data in stock_data_by_ticker.items():
            df = pd.DataFrame(data)
            df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
            stock_dataframes[ticker] = df

        print(f"    Loaded stock data for {len(stock_dataframes)} tickers")

        print("  Loading related companies...")
        validated_mentions = defaultdict(Counter)
        all_related_companies = defaultdict(list)

        related_sheet = wb["Related Companies"]
        related_rows = list(related_sheet.iter_rows(min_row=2, values_only=True))

        for row in related_rows:
            if len(row) >= 4:
                primary_ticker, related_ticker, count, rank = row[:4]
                if primary_ticker and related_ticker:
                    validated_mentions[primary_ticker][f"({related_ticker})"] = count

                    # Build all_related_companies (top 5)
                    if rank <= 5:
                        if related_ticker not in all_related_companies[primary_ticker]:
                            all_related_companies[primary_ticker].append(related_ticker)

        print(f"    Loaded related companies for {len(all_related_companies)} tickers")

        print("  Loading alpha formulas...")
        alpha_texts = {}

        alpha_sheet = wb["Alpha Formulas"]
        alpha_rows = list(alpha_sheet.iter_rows(min_row=2, values_only=True))

        for row in alpha_rows:
            if len(row) >= 3:
                ticker, related_companies, alpha_text = row[:3]
                if ticker and alpha_text:
                    alpha_texts[ticker] = alpha_text

        print(f"    ✓ Loaded alpha formulas for {len(alpha_texts)} tickers")

        wb.close()

        print(f"\n✓ All data loaded successfully")
        print(f"{'='*80}\n")

        return (daily_sentiments, stock_dataframes, validated_mentions,
                all_related_companies, alpha_texts, True)

    except Exception as e:
        print(f"❌ Error loading data: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None, None, None, False


def check_data_completeness(filename, required_tickers):
    """
    Check if Excel file has complete data for all required tickers

    Args:
        filename: Excel file path
        required_tickers: list of ticker symbols

    Returns:
        bool: True if data is complete, False otherwise
    """

    if not os.path.exists(filename):
        return False

    try:
        wb = load_workbook(filename, read_only=True)

        if "Metadata" not in wb.sheetnames:
            wb.close()
            return False

        meta_sheet = wb["Metadata"]
        meta_rows = list(meta_sheet.iter_rows(min_row=2, values_only=True))

        for row in meta_rows:
            if len(row) >= 2 and row[0] == "Data Complete":
                data_complete = row[1] == "Yes"
                wb.close()

                if not data_complete:
                    return False

        if "Stock Prices" not in wb.sheetnames:
            wb.close()
            return False

        stock_sheet = wb["Stock Prices"]
        stock_rows = list(stock_sheet.iter_rows(min_row=2, values_only=True))

        tickers_in_file = set()
        for row in stock_rows:
            if len(row) >= 1 and row[0]:
                tickers_in_file.add(row[0])

        wb.close()

        missing_tickers = set(required_tickers) - tickers_in_file

        if missing_tickers:
            print(f"Missing data for tickers: {', '.join(missing_tickers)}")
            return False

        return True

    except Exception as e:
        print(f"Error checking data completeness: {e}")
        return False


def display_data_summary(daily_sentiments, stock_dataframes, all_related_companies):
    """
    Display a summary of loaded data
    """
    print(f"\n{'='*80}")
    print("DATA SUMMARY")
    print(f"{'='*80}")

    print(f"\nSentiment Data:")
    for ticker in sorted(daily_sentiments.keys())[:10]:
        days = len(daily_sentiments[ticker])
        print(f"  {ticker}: {days} days")
    if len(daily_sentiments) > 10:
        print(f"  ... and {len(daily_sentiments) - 10} more tickers")

    print(f"\nStock Price Data:")
    for ticker in sorted(stock_dataframes.keys()):
        rows = len(stock_dataframes[ticker])
        print(f"  {ticker}: {rows} days")

    print(f"\nRelated Companies:")
    for ticker in sorted(all_related_companies.keys()):
        related = all_related_companies[ticker]
        print(f"  {ticker}: {', '.join(related)}")

    print(f"\n{'='*80}\n")