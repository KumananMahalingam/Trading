import time
import secret_key
from datetime import datetime, timezone
import spacy
from polygon import RESTClient
from collections import defaultdict, Counter
import json
import re
import pandas as pd
import yfinance as yf
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from groq import Groq
import torch
import os
import numpy as np
from lstm_predictor import train_stock_predictor, train_ensemble
from data_storage import (
    save_all_data_to_excel,
    load_all_data_from_excel,
)
from alternative_data_sources import (
    fetch_fred_data,
    fetch_all_alternative_data
)

client = RESTClient(secret_key.API_KEY)
groq_client = Groq(api_key=secret_key.GROQ_API_KEY)
nlp = spacy.load("en_core_web_sm")
sia = SentimentIntensityAnalyzer()

companies = {
    "AAPL": "Apple",
#    "JPM": "JPMorgan Chase & Co",
#    "PEP": "Pepsi",
#    "TM": "Toyota",
#    "AMZN": "Amazon"
}

COMPANY_ALIASES = {
    "APPLE": ["AAPL"],
#    "JPMORGAN": ["JPM", "JPM.N"],
#    "PEPSI": ["PEP", "PEPSICO"],
#    "TOYOTA": ["TM", "TYO"],
#    "AMAZON": ["AMZN"]
}

DATA_FILE = "stock_analysis_complete.xlsx"
FORCE_REFETCH = False  # Set to True to force re-fetching all data
FORCE_REGENERATE_ALPHAS = True

def load_company_tickers_json(file_path="company_tickers.json"):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"Loaded {len(data)} companies from {file_path}")
        ticker_to_info = {}
        name_to_ticker = {}
        name_variations = {}
        for key, company in data.items():
            ticker = company.get('ticker', '').upper()
            name = company.get('title', '')
            cik = company.get('cik_str', '')
            if ticker and name:
                ticker_to_info[ticker] = {'name': name, 'ticker': ticker, 'cik': cik}
                name_to_ticker[name.upper()] = ticker
                variations = generate_name_variations(name)
                for variation in variations:
                    if len(variation) >= 3:
                        name_variations[variation.upper()] = ticker
        print(f"Created {len(name_variations)} name variations for matching")
        return {
            'ticker_to_info': ticker_to_info,
            'name_to_ticker': name_to_ticker,
            'name_variations': name_variations
        }
    except FileNotFoundError:
        print(f"Error: {file_path} not found.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON: {e}")
        return None

def generate_name_variations(company_name):
    variations = [company_name]
    suffixes_pattern = r'\s+(Inc\.?|Corporation|Corp\.?|Company|Co\.?|Limited|Ltd\.?|Holdings|Holding|LLC|L\.L\.C\.|plc|PLC)'
    base_name = re.sub(suffixes_pattern, '', company_name, flags=re.IGNORECASE).strip()
    variations.append(base_name)
    variations.extend([
        company_name.replace(',', '').strip(),
        company_name.replace('.', '').strip(),
        base_name.replace(',', '').strip(),
        base_name.replace('.', '').strip(),
        company_name.replace('&', 'and'),
        company_name.replace(' and ', ' & '),
        base_name.replace('&', 'and'),
        base_name.replace(' and ', ' & ')
    ])
    words = base_name.split()
    if len(words) > 1:
        acronym = ''.join([word[0] for word in words if word[0].isupper()])
        if len(acronym) >= 2:
            variations.append(acronym)
    return list(set([v.strip() for v in variations if v.strip()]))

def validate_company_exists(company_name, company_lookups):
    if not company_lookups:
        return False, None, None
    company_upper = company_name.upper().strip()
    if company_upper in company_lookups['name_to_ticker']:
        ticker = company_lookups['name_to_ticker'][company_upper]
        official_name = company_lookups['ticker_to_info'][ticker]['name']
        return True, ticker, official_name
    if company_upper in company_lookups['name_variations']:
        ticker = company_lookups['name_variations'][company_upper]
        official_name = company_lookups['ticker_to_info'][ticker]['name']
        return True, ticker, official_name
    if company_upper in company_lookups['ticker_to_info']:
        ticker = company_upper
        official_name = company_lookups['ticker_to_info'][ticker]['name']
        return True, ticker, official_name
    for variation, ticker in company_lookups['name_variations'].items():
        if len(company_upper) >= 4 and company_upper in variation:
            official_name = company_lookups['ticker_to_info'][ticker]['name']
            return True, ticker, official_name
        elif len(variation) >= 4 and variation in company_upper:
            official_name = company_lookups['ticker_to_info'][ticker]['name']
            return True, ticker, official_name
    return False, None, None

def is_same_company(ticker1, ticker2, name1, name2):
    if ticker1 == ticker2:
        return True
    for base_name, aliases in COMPANY_ALIASES.items():
        if ticker1 in aliases and ticker2 in aliases:
            return True
    name1_base = re.sub(r'\s+(Inc\.?|Corporation|Corp\.?|Company|Co\.?|Limited|Ltd\.?)', '', name1, flags=re.IGNORECASE).strip().upper()
    name2_base = re.sub(r'\s+(Inc\.?|Corporation|Corp\.?|Company|Co\.?|Limited|Ltd\.?)', '', name2, flags=re.IGNORECASE).strip().upper()
    if name1_base == name2_base:
        return True
    return False

def validate_ticker_quality(ticker, min_market_cap=5e9, required_exchanges=['NYSE', 'NASDAQ', 'NYQ', 'NMS']):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        if not info or len(info) < 5:
            return False, "Insufficient data"
        exchange = info.get('exchange', '').upper()
        if exchange not in required_exchanges:
            return False, f"Not on major exchange (found: {exchange})"
        market_cap = info.get('marketCap', 0)
        if market_cap < min_market_cap:
            return False, f"Market cap too low (${market_cap:,.0f})"
        volume = info.get('volume', 0)
        if volume < 100000:
            return False, "Low trading volume"
        return True, "Valid"
    except Exception as e:
        return False, f"Error: {str(e)}"

def fetch_stock_data(ticker, start_date, end_date):
    try:
        print(f"  Fetching stock data for {ticker}...")
        start = start_date.split('T')[0]
        end = end_date.split('T')[0]
        stock = yf.Ticker(ticker)
        df = stock.history(start=start, end=end)
        if df.empty:
            print(f"  No data available for {ticker}")
            return pd.DataFrame()
        df.reset_index(inplace=True)
        df.columns = df.columns.str.lower()
        df = df.rename(columns={'date': 'date'})
        df['date'] = df['date'].dt.strftime('%Y-%m-%d')
        df = df[['date', 'open', 'high', 'low', 'close', 'volume']]
        print(f"  Fetched {len(df)} days of stock data for {ticker}")
        return df
    except Exception as e:
        print(f"  Error fetching stock data for {ticker}: {e}")
        return pd.DataFrame()

def calculate_adx(df, period=14):
    """Calculate Average Directional Index (ADX)"""
    high_diff = df['high'].diff()
    low_diff = -df['low'].diff()

    plus_dm = high_diff.where((high_diff > low_diff) & (high_diff > 0), 0)
    minus_dm = low_diff.where((low_diff > high_diff) & (low_diff > 0), 0)

    high_low = df['high'] - df['low']
    high_close = np.abs(df['high'] - df['close'].shift())
    low_close = np.abs(df['low'] - df['close'].shift())
    ranges = pd.concat([high_low, high_close, low_close], axis=1)
    atr = np.max(ranges, axis=1).rolling(period).mean()

    plus_di = 100 * (plus_dm.rolling(period).mean() / atr)
    minus_di = 100 * (minus_dm.rolling(period).mean() / atr)

    dx = 100 * np.abs(plus_di - minus_di) / (plus_di + minus_di)
    adx = dx.rolling(period).mean()

    return adx

def calculate_technical_indicators(df):
    """ENHANCED: Calculate comprehensive technical indicators"""
    if df.empty:
        return df

    df['SMA_5'] = df['close'].rolling(window=5).mean()
    df['SMA_10'] = df['close'].rolling(window=10).mean()
    df['SMA_20'] = df['close'].rolling(window=20).mean()
    df['SMA_50'] = df['close'].rolling(window=50).mean()
    df['SMA_200'] = df['close'].rolling(window=200).mean()

    df['EMA_12'] = df['close'].ewm(span=12, adjust=False).mean()
    df['EMA_26'] = df['close'].ewm(span=26, adjust=False).mean()

    df['MACD'] = df['EMA_12'] - df['EMA_26']
    df['MACD_Signal'] = df['MACD'].ewm(span=9, adjust=False).mean()
    df['MACD_Hist'] = df['MACD'] - df['MACD_Signal']

    delta = df['close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    df['RSI'] = 100 - (100 / (1 + rs))

    df['BB_Middle'] = df['close'].rolling(window=20).mean()
    bb_std = df['close'].rolling(window=20).std()
    df['BB_Upper'] = df['BB_Middle'] + (bb_std * 2)
    df['BB_Lower'] = df['BB_Middle'] - (bb_std * 2)
    df['BB_Width'] = (df['BB_Upper'] - df['BB_Lower']) / df['BB_Middle']
    df['BB_Position'] = (df['close'] - df['BB_Lower']) / (df['BB_Upper'] - df['BB_Lower'])

    df['Return_1D'] = df['close'].pct_change(1)
    df['Return_5D'] = df['close'].pct_change(5)
    df['Return_20D'] = df['close'].pct_change(20)
    df['Return_60D'] = df['close'].pct_change(60)

    df['Volatility_5D'] = df['Return_1D'].rolling(window=5).std()
    df['Volatility_20D'] = df['Return_1D'].rolling(window=20).std()
    df['Volatility_60D'] = df['Return_1D'].rolling(window=60).std()

    df['Volume_SMA_20'] = df['volume'].rolling(window=20).mean()
    df['Volume_Ratio'] = df['volume'] / df['Volume_SMA_20']

    df['Momentum_5'] = df['close'] - df['close'].shift(5)
    df['Momentum_10'] = df['close'] - df['close'].shift(10)
    df['Momentum_20'] = df['close'] - df['close'].shift(20)

    df['ROC_5'] = ((df['close'] - df['close'].shift(5)) / df['close'].shift(5)) * 100
    df['ROC_10'] = ((df['close'] - df['close'].shift(10)) / df['close'].shift(10)) * 100

    # Average True Range (ATR)
    high_low = df['high'] - df['low']
    high_close = np.abs(df['high'] - df['close'].shift())
    low_close = np.abs(df['low'] - df['close'].shift())
    ranges = pd.concat([high_low, high_close, low_close], axis=1)
    true_range = np.max(ranges, axis=1)
    df['ATR'] = true_range.rolling(14).mean()

    # On-Balance Volume (OBV)
    df['OBV'] = (np.sign(df['close'].diff()) * df['volume']).fillna(0).cumsum()

    # Money Flow Index (MFI)
    typical_price = (df['high'] + df['low'] + df['close']) / 3
    money_flow = typical_price * df['volume']

    positive_flow = money_flow.where(typical_price > typical_price.shift(1), 0).rolling(14).sum()
    negative_flow = money_flow.where(typical_price < typical_price.shift(1), 0).rolling(14).sum()
    mfi_ratio = positive_flow / negative_flow
    df['MFI'] = 100 - (100 / (1 + mfi_ratio))

    # Stochastic Oscillator
    low_14 = df['low'].rolling(14).min()
    high_14 = df['high'].rolling(14).max()
    df['Stochastic_K'] = 100 * (df['close'] - low_14) / (high_14 - low_14)
    df['Stochastic_D'] = df['Stochastic_K'].rolling(3).mean()

    # Williams R
    df['Williams_R'] = -100 * (high_14 - df['close']) / (high_14 - low_14)

    # Trend strength (ADX)
    df['ADX'] = calculate_adx(df)

    return df

def prepare_dataframe_for_alpha(ticker, stock_df, daily_sentiments, related_companies, alternative_data=None, economic_data=None):
    """ENHANCED: Prepare dataframe with sentiment lag features"""
    if stock_df.empty:
        return None

    df = calculate_technical_indicators(stock_df.copy())

    sentiment_dict = {}
    for date, scores in daily_sentiments[ticker].items():
        sentiment_dict[date] = sum(scores) / len(scores) if scores else 0
    df[f'{ticker}_Sentiment'] = df['date'].map(sentiment_dict).fillna(0)

    df[f'{ticker}_Sentiment_Lag1'] = df[f'{ticker}_Sentiment'].shift(1)
    df[f'{ticker}_Sentiment_Lag5'] = df[f'{ticker}_Sentiment'].shift(5)

    df[f'{ticker}_Sentiment_Change'] = df[f'{ticker}_Sentiment'].diff()
    df[f'{ticker}_Sentiment_MA5'] = df[f'{ticker}_Sentiment'].rolling(5).mean()

    for related_ticker in related_companies:
        if related_ticker in daily_sentiments:
            related_sentiment_dict = {}
            for date, scores in daily_sentiments[related_ticker].items():
                related_sentiment_dict[date] = sum(scores) / len(scores) if scores else 0
            df[f'{related_ticker}_Sentiment'] = df['date'].map(related_sentiment_dict).fillna(0)

            df[f'Sentiment_Div_{ticker}_{related_ticker}'] = \
                df[f'{ticker}_Sentiment'] - df[f'{related_ticker}_Sentiment']

    if alternative_data and ticker in alternative_data:
        from alternative_data_sources import add_alternative_features_to_df
        df = add_alternative_features_to_df(
            df, ticker,
            economic_data if economic_data is not None else pd.DataFrame(),
            alternative_data[ticker]['sec'],
            alternative_data[ticker]['earnings']
        )

    return df

def generate_simple_alphas(ticker):
    """Fallback alphas that don't require related company sentiment"""
    return f"""
α1 = Return_5D
α2 = (RSI - 50) / 50
α3 = MACD - MACD_Signal
α4 = (close - SMA_20) / SMA_20
α5 = Return_5D + 0.1 * {ticker}_Sentiment
"""

def generate_alphas_with_groq(ticker, company_name, comprehensive_df, related_companies, max_retries=3):
    """
    Generate predictive alphas using Groq LLM with enhanced prompt
    """
    print(f"\n  Generating alphas for {company_name} ({ticker}) using Groq LLM...")

    cutoff_idx = int(len(comprehensive_df) * 0.8)
    sample_df = comprehensive_df.iloc[:cutoff_idx].tail(20).copy()  # Reduced to 20 days to fit token limit

    # Select only most important columns to reduce token usage
    important_cols = ['date', 'close', 'volume']

    tech_cols = [col for col in sample_df.columns if any(ind in col for ind in
                ['RSI', 'MACD', 'SMA_20', 'SMA_50', 'EMA_12', 'EMA_26', 'ATR',
                 'BB_Upper', 'BB_Lower', 'Return_5D', 'Return_20D', 'Volatility_20D', 'Volume_Ratio'])]
    important_cols.extend(tech_cols)

    sentiment_cols = [col for col in sample_df.columns if 'Sentiment' in col]
    important_cols.extend(sentiment_cols)

    econ_cols = [col for col in sample_df.columns if any(ind in col for ind in
                ['VIX', 'Fed_Funds_Rate', 'Treasury_10Y', 'CPI', 'Unemployment_Rate'])]
    important_cols.extend(econ_cols)

    alt_cols = [col for col in sample_df.columns if any(ind in col for ind in
               ['SEC_Sentiment', 'Earnings', 'Days_Since_Earnings'])]
    important_cols.extend(alt_cols)

    important_cols = list(dict.fromkeys(important_cols))  # Remove duplicates while preserving order
    sample_df = sample_df[[col for col in important_cols if col in sample_df.columns]]

    available_features = list(sample_df.columns)

    # Categorize features with counts
    stock_features = [col for col in available_features if col in ['close', 'open', 'high', 'low', 'volume', 'date']]

    technical_indicators = [col for col in available_features if any(ind in col.upper() for ind in
                           ['SMA', 'EMA', 'RSI', 'MACD', 'BB_', 'ATR', 'OBV', 'MFI', 'STOCHASTIC',
                            'WILLIAMS', 'ADX', 'ROC', 'MOMENTUM', 'VOLATILITY', 'RETURN'])]

    sentiment_features = [col for col in available_features if 'Sentiment' in col]

    target_sentiment = [col for col in sentiment_features if col.startswith(f'{ticker}_')]
    related_sentiment = [col for col in sentiment_features if not col.startswith(f'{ticker}_') and not col.startswith('Sentiment_Div')]
    sentiment_divergence = [col for col in sentiment_features if col.startswith('Sentiment_Div')]

    economic_features = [col for col in available_features if any(ind in col for ind in
                        ['GDP', 'CPI', 'Unemployment', 'Fed_Funds', 'Treasury', 'VIX', 'Oil_Price',
                         'Gold_Price', 'Consumer_Sentiment', 'Retail_Sales'])]

    alternative_features = [col for col in available_features if any(ind in col for ind in
                           ['SEC', 'Earnings', 'Days_Since_Earnings'])]

    target_col = ticker + '_Sentiment'
    stats_context = ""
    if target_col in sample_df.columns:
        sent_mean = sample_df[target_col].mean()
        sent_std = sample_df[target_col].std()
        sent_min = sample_df[target_col].min()
        sent_max = sample_df[target_col].max()

        if 'Volatility_20D' in sample_df.columns:
            vol_20d_val = sample_df['Volatility_20D'].iloc[-1]
            vol_20d_str = f"{vol_20d_val:.4f}"
        else:
            vol_20d_str = 'N/A'

        if 'Return_20D' in sample_df.columns:
            ret_20d_val = sample_df['Return_20D'].iloc[-1]
            ret_20d_str = f"{ret_20d_val:.4f}"
        else:
            ret_20d_str = 'N/A'

        stats_context = f"""
Statistical Context:
- {ticker} Sentiment: Mean={sent_mean:.3f}, Std={sent_std:.3f}, Range=[{sent_min:.3f}, {sent_max:.3f}]
- Recent volatility (20D): {vol_20d_str}
- Recent return (20D): {ret_20d_str}
"""

    related_str = "None available"
    if related_companies:
        related_with_overlap = []
        for rel_ticker in related_companies:
            if f'{rel_ticker}_Sentiment' in sample_df.columns:
                related_with_overlap.append(f"{rel_ticker} (sentiment available)")
            else:
                related_with_overlap.append(f"{rel_ticker} (no sentiment)")
        related_str = ", ".join(related_with_overlap)

    df_summary = f"""
Date Range: {sample_df['date'].iloc[0]} to {sample_df['date'].iloc[-1]} ({len(sample_df)} days)

Recent Values (last 5 days):
{sample_df.tail(5).to_string(index=False, max_cols=10)}

Feature Statistics:
{sample_df.describe().loc[['mean', 'std', 'min', 'max']].to_string()}
"""

    example_alpha_1 = f"Sentiment_Div_{ticker}_MSFT / (Volatility_20D + 1e-8)"
    example_alpha_2 = f"(RSI - 50) / 50 + 0.3 * {ticker}_Sentiment_MA5"
    example_alpha_3 = f"(MACD_Hist / ATR) * (1 + {ticker}_Sentiment_Change)"
    example_alpha_4 = f"(SMA_5 - SMA_50) / close * (1 - Unemployment_Rate / 10)"
    example_alpha_5 = f"Return_20D * Volume_Ratio * (1 + {ticker}_Days_Since_Earnings / 90)"

    sentiment_example_features = f"{ticker}_Sentiment_MA5, {ticker}_Sentiment_Lag5"

    prompt = f"""You are a quantitative finance expert designing predictive alpha factors for algorithmic trading.

TASK: Generate 5 formulaic alpha signals to predict {company_name} ({ticker}) stock price movements.

DATA AVAILABLE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. STOCK FEATURES ({len(stock_features)}):
   {', '.join(stock_features)}

2. TECHNICAL INDICATORS ({len(technical_indicators)}):
   Available: RSI, MACD, SMA (5/10/20/50/200), EMA (12/26), BB, ATR, Volume_Ratio, Returns, Volatility

3. SENTIMENT FEATURES ({len(sentiment_features)}):
   Target: {', '.join(target_sentiment[:5])}
   Related: {', '.join(related_sentiment[:5])}{'...' if len(related_sentiment) > 5 else ''}
   Divergences: {len(sentiment_divergence)} available

4. ECONOMIC INDICATORS ({len(economic_features)}):
   Available: {', '.join(economic_features[:10])}{'...' if len(economic_features) > 10 else ''}

5. ALTERNATIVE DATA ({len(alternative_features)}):
   {', '.join(alternative_features)}

RELATED COMPANIES: {related_str}

{stats_context}

SAMPLE DATA (Last 20 trading days - summary):
{df_summary}

ALPHA GENERATION REQUIREMENTS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Generate exactly 5 diverse alpha formulas following these guidelines:

1. UTILIZE MULTIPLE DATA SOURCES:
   - α1: Focus on CROSS-COMPANY SENTIMENT DIVERGENCE (use Sentiment_Div features or create divergences)
   - α2: Combine TECHNICAL INDICATORS with TARGET SENTIMENT
   - α3: Use MOMENTUM + VOLATILITY signals (trend-following)
   - α4: Incorporate ECONOMIC INDICATORS + stock technicals
   - α5: Alternative data fusion (SEC sentiment, earnings events, or multi-timeframe technical)

2. FEATURE ENGINEERING BEST PRACTICES:
   ✓ Normalize by volatility (divide by ATR or Volatility_20D)
   ✓ Use ratio features (e.g., Volume_Ratio, BB_Position)
   ✓ Combine short-term and long-term signals (e.g., SMA_5 - SMA_50)
   ✓ Leverage sentiment lags and moving averages (e.g., {sentiment_example_features})
   ✓ Create divergence signals between related metrics

3. MATHEMATICAL VALIDITY:
   ✓ Avoid division by zero (use ATR, std, or add small epsilon: 1e-8)
   ✓ Keep formulas executable (2-6 terms per alpha)
   ✓ Use ONLY column names from the data above
   ✓ Ensure formulas return numeric values

4. STRATEGIC DIVERSITY:
   - Include at least ONE mean-reversion alpha (e.g., Bollinger Band position)
   - Include at least ONE momentum/trend alpha (e.g., MACD, ROC)
   - Include at least ONE sentiment-based alpha
   - Avoid redundant alphas (don't repeat similar logic)

OUTPUT FORMAT (respond with ONLY these 5 lines, no explanations):
α1 = <formula>
α2 = <formula>
α3 = <formula>
α4 = <formula>
α5 = <formula>

EXAMPLE ALPHAS (for reference only - generate NEW ones based on actual data):
α1 = {example_alpha_1}
α2 = {example_alpha_2}
α3 = {example_alpha_3}
α4 = {example_alpha_4}
α5 = {example_alpha_5}

NOW GENERATE 5 UNIQUE ALPHAS FOR {company_name}:"""

    for attempt in range(max_retries):
        try:
            print(f"    Attempt {attempt + 1}/{max_retries}...")

            response = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {
                        "role": "system",
                        "content": "You are a quantitative finance expert specializing in alpha generation. Output ONLY the 5 alpha formulas in the exact format specified. No preamble, no explanations, no markdown formatting."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                temperature=0.8,
                max_tokens=1500,
                top_p=0.95
            )

            alpha_text = response.choices[0].message.content.strip()

            # Extract alpha lines
            alpha_lines = [line for line in alpha_text.split('\n') if line.strip().startswith('α')]

            if len(alpha_lines) >= 5:
                print(f"    ✓ Successfully generated {len(alpha_lines)} alphas")

                all_alphas_text = '\n'.join(alpha_lines)
                features_used = set()

                for feature_group in [sentiment_divergence, target_sentiment,
                                     technical_indicators[:20], economic_features[:10]]:
                    for feature in feature_group:
                        if feature in all_alphas_text:
                            features_used.add(feature)

                print(f"    ✓ Alphas utilize {len(features_used)} unique features")

                if any('Sentiment_Div' in line for line in alpha_lines):
                    print(f"    ✓ Sentiment divergence signals included")

                return alpha_text
            else:
                print(f"    ⚠ Only got {len(alpha_lines)} alphas, retrying...")

        except Exception as e:
            print(f"    ✗ Error on attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(2)
            else:
                print(f"    Falling back to simple alphas for {ticker}")
                return generate_simple_alphas(ticker)

    return generate_simple_alphas(ticker)

def fetch_news(ticker, start_date, end_date, batch_size=1000, sleep_time=12):
    results = []
    batch_count = 0
    try:
        response = client.list_ticker_news(
            ticker=ticker,
            published_utc_gte=start_date,
            published_utc_lte=end_date,
            limit=batch_size,
            order="asc"
        )
        current_batch = []
        for item in response:
            current_batch.append(item)
            if len(current_batch) >= batch_size:
                break
        while current_batch:
            results.extend(current_batch)
            batch_count += 1
            print(f"  Fetched batch {batch_count} ({len(current_batch)} articles) for {ticker}. Total: {len(results)}")
            if len(current_batch) < batch_size:
                print(f"  Reached end of articles for {ticker}")
                break
            last_article_date = current_batch[-1].published_utc
            if last_article_date >= end_date:
                print(f"  Reached end date for {ticker}")
                break
            print(f"  Waiting {sleep_time} seconds before next batch...")
            time.sleep(sleep_time)
            response = client.list_ticker_news(
                ticker=ticker,
                published_utc_gt=last_article_date,
                published_utc_lte=end_date,
                limit=batch_size,
                order='asc'
            )
            current_batch = []
            for item in response:
                current_batch.append(item)
                if len(current_batch) >= batch_size:
                    break
        print(f"  Total articles fetched for {ticker}: {len(results)}")
    except Exception as e:
        print(f"Error fetching news for {ticker}: {e}")
    return results

def fetch_sentiment_for_ticker(ticker, start_date, end_date, daily_sentiments):
    """Fetch news and compute sentiment for a specific ticker"""
    print(f"  Fetching sentiment data for {ticker}...")

    news_articles = fetch_news(ticker, start_date, end_date, batch_size=1000, sleep_time=12)

    if not news_articles:
        print(f"    No news articles found for {ticker}")
        return

    for item in news_articles:
        text = f"{item.title} {item.summary if hasattr(item, 'summary') else ''}"
        sentiment_scores = sia.polarity_scores(text)
        compound_score = sentiment_scores['compound']

        article_date = item.published_utc.split('T')[0] if 'T' in item.published_utc else item.published_utc[:10]
        daily_sentiments[ticker][article_date].append(compound_score)

    print(f"    Processed {len(news_articles)} articles, {len(daily_sentiments[ticker])} days of sentiment")

def check_essential_data_only(file_path, tickers):
    """Check only for daily sentiments, stock data, and related companies"""
    try:
        if not os.path.exists(file_path):
            return False

        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames

        if 'Daily Sentiment' not in sheet_names:
            print("  Missing: Daily Sentiment sheet")
            wb.close()
            return False

        if 'Related Companies' not in sheet_names:
            print("  Missing: Related Companies sheet")
            wb.close()
            return False

        if 'Stock Prices' not in sheet_names:
            print("  Missing: Stock Prices sheet")
            wb.close()
            return False

        sentiments_sheet = wb['Daily Sentiment']
        related_sheet = wb['Related Companies']
        stock_sheet = wb['Stock Prices']

        sentiment_rows = list(sentiments_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        related_rows = list(related_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        stock_rows = list(stock_sheet.iter_rows(min_row=2, values_only=True))

        if not sentiment_rows or not sentiment_rows[0][0]:
            print("  Daily Sentiment sheet is empty")
            wb.close()
            return False

        if not related_rows or not related_rows[0][0]:
            print("  Related Companies sheet is empty")
            wb.close()
            return False

        if not stock_rows or not stock_rows[0][0]:
            print("  Stock Prices sheet is empty")
            wb.close()
            return False

        tickers_in_file = set()
        for row in stock_rows:
            if len(row) >= 1 and row[0]:
                tickers_in_file.add(row[0])

        wb.close()

        missing_tickers = set(tickers) - tickers_in_file
        if missing_tickers:
            print(f"  Missing stock data for: {', '.join(missing_tickers)}")
            return False

        print("  ✓ All essential data found in cache")
        return True

    except Exception as e:
        print(f"  Error checking cache: {e}")
        import traceback
        traceback.print_exc()
        return False

print("="*80)
print("IMPROVED STOCK PRICE PREDICTION PIPELINE")
print("="*80)

print("\nLoading company tickers JSON...")
company_lookups = load_company_tickers_json("company_tickers.json")

load_from_cache = False

if not FORCE_REFETCH and check_essential_data_only(DATA_FILE, list(companies.keys())):
    print(f"\n{'='*80}")
    print("COMPLETE DATA FOUND IN CACHE")
    print(f"{'='*80}")

    user_input = input("\nLoad data from cache? (yes/no) [yes]: ").strip().lower()
    if user_input in ['', 'yes', 'y']:
        load_from_cache = True

if load_from_cache:
    (daily_sentiments, stock_dataframes, validated_mentions,
     all_related_companies, alpha_texts, success) = load_all_data_from_excel(DATA_FILE)

    if not success:
        print("Failed to load data from cache. Falling back to data collection...")
        load_from_cache = False
    else:
        print("\n✓ Loaded essential data from cache")
        print(f"  - Daily sentiment for {len(daily_sentiments)} tickers")
        print(f"  - Stock data for {len(stock_dataframes)} tickers")
        print(f"  - Related companies for {len(all_related_companies)} tickers")
        print(f"  - Alpha formulas for {len(alpha_texts)} tickers")

        print("\n" + "="*80)
        print("FETCHING ECONOMIC & FUNDAMENTAL DATA (FRED + SEC + EARNINGS)")
        print("="*80)

        all_dates = []
        for ticker, df in stock_dataframes.items():
            if not df.empty and 'date' in df.columns:
                all_dates.extend(df['date'].tolist())

        if all_dates:
            full_start_date = min(all_dates)
            full_end_date = max(all_dates)
            print(f"Date range from cached data: {full_start_date} to {full_end_date}")

            print("\nFetching FRED economic indicators...")
            economic_data = fetch_fred_data(full_start_date, full_end_date)

            print("\nFetching SEC filings and earnings data...")
            all_alternative_data = {}

            for ticker, name in companies.items():
                print(f"\n{ticker} ({name}):")

                try:
                    sec_sentiment, earnings_sentiment = fetch_all_alternative_data(
                        ticker,
                        full_start_date,
                        full_end_date
                    )

                    all_alternative_data[ticker] = {
                        'sec': sec_sentiment,
                        'earnings': earnings_sentiment
                    }

                    time.sleep(3)
                except Exception as e:
                    print(f"  ✗ Error fetching alternative data for {ticker}: {e}")
                    all_alternative_data[ticker] = {
                        'sec': {},
                        'earnings': {}
                    }

            print("\n✓ Alternative data collection complete")
        else:
            print("\n⚠ Could not determine date range from cached data")
            print("  Using empty economic data")
            economic_data = pd.DataFrame()
            all_alternative_data = {}

        print("\n" + "="*80)
        print("CHECKING ALPHA FORMULAS")
        print("="*80)

        missing_alphas = []
        for ticker in companies.keys():
            if FORCE_REGENERATE_ALPHAS or ticker not in alpha_texts or not alpha_texts[ticker]:
                missing_alphas.append(ticker)
                if FORCE_REGENERATE_ALPHAS:
                    print(f"  🔄 Forcing regeneration for {ticker}")
            else:
                print(f"  ✓ Loaded alpha formula for {ticker}")

        if missing_alphas:
            print(f"\n{'='*80}")
            print(f"GENERATING MISSING ALPHAS FOR {len(missing_alphas)} TICKERS")
            print(f"{'='*80}")

            for ticker in missing_alphas:
                name = companies[ticker]

                if ticker not in stock_dataframes:
                    print(f"\nSkipping {ticker}: No stock data available")
                    alpha_texts[ticker] = generate_simple_alphas(ticker)
                    continue

                print(f"\nGenerating alphas for {name} ({ticker})...")

                stock_df = stock_dataframes[ticker]
                related_companies = all_related_companies.get(ticker, [])

                comprehensive_df = prepare_dataframe_for_alpha(
                    ticker,
                    stock_df,
                    daily_sentiments,
                    related_companies,
                    alternative_data=all_alternative_data,
                    economic_data=economic_data
                )

                if comprehensive_df is None or comprehensive_df.empty:
                    print(f"  Failed to prepare data, using simple alphas")
                    alpha_texts[ticker] = generate_simple_alphas(ticker)
                    continue

                alpha_text = generate_alphas_with_groq(
                    ticker,
                    name,
                    comprehensive_df,
                    related_companies
                )

                alpha_texts[ticker] = alpha_text

                print(f"\nGenerated alphas for {ticker}:")
                print(alpha_text)
                print()

                time.sleep(2)

            save_all_data_to_excel(
                DATA_FILE,
                daily_sentiments,
                stock_dataframes,
                validated_mentions,
                all_related_companies,
                alpha_texts
            )
        else:
            print(f"\n✓ All alpha formulas loaded from cache - no API calls needed!")

if not load_from_cache:

    start_date = datetime(2023, 9, 1, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
    end_date   = datetime(2025, 9, 1, tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

    daily_sentiments = defaultdict(lambda: defaultdict(list))
    validated_mentions = defaultdict(Counter)
    stock_dataframes = {}
    alpha_texts = {}
    all_related_companies = {}

    print("\n" + "="*80)
    print("PHASE 1: FETCHING NEWS FOR PRIMARY COMPANIES")
    print("="*80)

    for ticker, name in companies.items():
        print(f"\nFetching news for {name} ({ticker})...")
        news_articles = fetch_news(ticker, start_date, end_date, batch_size=1000, sleep_time=12)

        for item in news_articles:
            text = f"{item.title} {item.summary if hasattr(item, 'summary') else ''}"
            sentiment_scores = sia.polarity_scores(text)
            compound_score = sentiment_scores['compound']

            article_date = item.published_utc.split('T')[0] if 'T' in item.published_utc else item.published_utc[:10]
            daily_sentiments[ticker][article_date].append(compound_score)

            doc = nlp(text)
            for ent in doc.ents:
                if ent.label_ == "ORG":
                    if company_lookups:
                        is_valid, mention_ticker, official_name = validate_company_exists(ent.text, company_lookups)
                        if is_valid:
                            validated_mentions[ticker][f"{official_name} ({mention_ticker})"] += 1

    print("\n" + "="*80)
    print("PHASE 2: IDENTIFYING AND FETCHING RELATED COMPANY SENTIMENT")
    print("="*80)

    all_related_tickers = set()

    for ticker, name in companies.items():
        print(f"\nIdentifying related companies for {name} ({ticker})...")

        related_companies = []
        for company_str, count in validated_mentions[ticker].most_common(20):
            if len(related_companies) >= 5:
                break

            match = re.search(r'\(([A-Z]+)\)', company_str)
            if not match:
                continue

            related_ticker = match.group(1)

            if related_ticker == ticker:
                continue

            related_name = company_str.split('(')[0].strip()

            if is_same_company(ticker, related_ticker, name, related_name):
                continue

            is_valid, reason = validate_ticker_quality(related_ticker, min_market_cap=5e9)
            if not is_valid:
                print(f"     Skipped {related_ticker}: {reason}")
                continue

            related_companies.append(related_ticker)
            all_related_tickers.add(related_ticker)
            print(f"     Added {related_ticker} ({related_name}): {count} mentions")

        all_related_companies[ticker] = related_companies
        print(f"  Found {len(related_companies)} related companies")

    print(f"\n{'='*80}")
    print(f"FETCHING SENTIMENT FOR {len(all_related_tickers)} UNIQUE RELATED COMPANIES")
    print(f"{'='*80}")

    for related_ticker in all_related_tickers:
        print(f"\nFetching sentiment for {related_ticker}...")
        fetch_sentiment_for_ticker(related_ticker, start_date, end_date, daily_sentiments)
        print(f"Waiting 15 seconds before next company...")
        time.sleep(15)

    print("\n" + "="*80)
    print("FETCHING ECONOMIC & FUNDAMENTAL DATA")
    print("="*80)

    full_start_date = start_date[:10]
    full_end_date = end_date[:10]

    from alternative_data_sources import fetch_fred_data, fetch_all_alternative_data

    economic_data = fetch_fred_data(full_start_date, full_end_date)

    all_alternative_data = {}

    for ticker, name in companies.items():
        print(f"\n{ticker} ({name}):")

        sec_sentiment, earnings_sentiment = fetch_all_alternative_data(
            ticker,
            full_start_date,
            full_end_date
        )

        all_alternative_data[ticker] = {
            'sec': sec_sentiment,
            'earnings': earnings_sentiment
        }

        time.sleep(3)

    print("\n Alternative data collection complete")

    print("\n" + "="*80)
    print("PHASE 3: FETCHING STOCK DATA")
    print("="*80)

    for ticker, name in companies.items():
        print(f"\nFetching stock data for {name} ({ticker})...")
        stock_df = fetch_stock_data(ticker, start_date, end_date)
        if not stock_df.empty:
            stock_dataframes[ticker] = stock_df
        time.sleep(2)

    print("\n" + "="*80)
    print("PHASE 3.5: GENERATING ALPHAS WITH GROQ LLM")
    print("="*80)

    for ticker, name in companies.items():
        if ticker not in stock_dataframes:
            print(f"\nSkipping {ticker}: No stock data available")
            alpha_texts[ticker] = generate_simple_alphas(ticker)
            continue

        print(f"\nGenerating alphas for {name} ({ticker})...")

        stock_df = stock_dataframes[ticker]
        related_companies = all_related_companies.get(ticker, [])

        comprehensive_df = prepare_dataframe_for_alpha(
            ticker,
            stock_df,
            daily_sentiments,
            related_companies,
            alternative_data=all_alternative_data,
            economic_data=economic_data
        )

        if comprehensive_df is None or comprehensive_df.empty:
            print(f"  Failed to prepare data, using simple alphas")
            alpha_texts[ticker] = generate_simple_alphas(ticker)
            continue

        alpha_text = generate_alphas_with_groq(
            ticker,
            name,
            comprehensive_df,
            related_companies
        )

        alpha_texts[ticker] = alpha_text

        print(f"\nGenerated alphas for {ticker}:")
        print(alpha_text)
        print()

    time.sleep(2)

    save_all_data_to_excel(
        DATA_FILE,
        daily_sentiments,
        stock_dataframes,
        validated_mentions,
        all_related_companies,
        alpha_texts
    )

print("\n" + "="*80)
print("PHASE 4: TRAINING ENSEMBLE OF ENHANCED LSTM MODELS")
print("="*80)

device = 'cuda' if torch.cuda.is_available() else 'cpu'
print(f"Using device: {device}")

trained_models = {}

successful_trainings = 0

for ticker, name in companies.items():
    if ticker not in stock_dataframes or ticker not in alpha_texts:
        print(f"\nSkipping {ticker}: Missing data or alphas")
        continue

    print(f"\n{'='*80}")
    print(f"TRAINING ENSEMBLE FOR {name} ({ticker})")
    print(f"{'='*80}")

    stock_df = stock_dataframes[ticker]
    related_companies = all_related_companies.get(ticker, [])

    comprehensive_df = prepare_dataframe_for_alpha(
        ticker,
        stock_df,
        daily_sentiments,
        related_companies,
        alternative_data=all_alternative_data,
        economic_data=economic_data
    )

    if comprehensive_df is None or comprehensive_df.empty:
        print(f" Failed to prepare data for {ticker}")
        continue

    try:
        alpha_text = alpha_texts.get(ticker, generate_simple_alphas(ticker))

        if len(comprehensive_df) < 100:
            print(f" Insufficient data for {ticker}: only {len(comprehensive_df)} rows")
            print(f"   Need at least 100 rows for training")
            continue

        print(f" Data ready: {len(comprehensive_df)} rows with {len(comprehensive_df.columns)} features")

        print(f"\nAttempting ensemble training...")
        ensemble, metrics, scalers = train_ensemble(
            ticker=ticker,
            company_name=name,
            comprehensive_df=comprehensive_df,
            alpha_text=alpha_text,
            n_models=2,  # Start with 2 models
            window_size=30,
            device=device
        )

        if ensemble is not None and metrics is not None:
            trained_models[ticker] = {
                'model': ensemble,
                'metrics': metrics,
                'scalers': scalers,
                'type': 'ensemble'
            }

            torch.save(ensemble.state_dict(), f'{ticker}_ensemble.pth')
            print(f"\n Ensemble saved as {ticker}_ensemble.pth")

            for i, model in enumerate(ensemble.models):
                torch.save(model.state_dict(), f'{ticker}_model_{i}.pth')

            successful_trainings += 1

        else:
            print(f" Ensemble training failed, trying single model...")

            model, metrics, scalers = train_stock_predictor(
                ticker=ticker,
                company_name=name,
                comprehensive_df=comprehensive_df,
                alpha_text=alpha_text,
                window_size=30,
                num_epochs=100,
                device=device
            )

            if model is not None and metrics is not None:
                trained_models[ticker] = {
                    'model': model,
                    'metrics': metrics,
                    'scalers': scalers,
                    'type': 'single'
                }
                torch.save(model.state_dict(), f'{ticker}_model.pth')
                print(f" Single model saved as {ticker}_model.pth")
                successful_trainings += 1
            else:
                print(f" Single model training also failed for {ticker}")

    except Exception as e:
        print(f"\n Error training model for {ticker}: {e}")
        import traceback
        traceback.print_exc()
        print(f"Continuing with next ticker...")
        continue

    print(f"\nWaiting 3 seconds before next ticker...")
    time.sleep(3)

print(f"\n{'='*80}")
print(f"TRAINING SUMMARY")
print(f"{'='*80}")
print(f"Successfully trained models: {successful_trainings}/{len(companies)}")
print(f"Models in trained_models dict: {len(trained_models)}")

if trained_models:
    print("\nUpdating Excel file with training results...")
    save_all_data_to_excel(
        DATA_FILE,
        daily_sentiments,
        stock_dataframes,
        validated_mentions,
        all_related_companies,
        alpha_texts,
        trained_models
    )

print("\n" + "="*80)
print("PIPELINE COMPLETE!")
print("="*80)
print(f"\nModels trained: {len(trained_models)}/{len(companies)}")

for ticker in trained_models:
    model_data = trained_models[ticker]
    model_type = model_data.get('type', 'single')
    metrics = model_data['metrics']

    print(f"\n  {ticker} ({model_type}):")
    print(f"    - RMSE: {metrics.get('RMSE', 0):.6f}")
    print(f"    - Directional Accuracy: {metrics.get('Directional Accuracy', 0):.2f}%")
    print(f"    - Up Precision: {metrics.get('Up Precision', 0):.2f}%")
    print(f"    - Down Precision: {metrics.get('Down Precision', 0):.2f}%")
    print(f"    - Large Move Hit Rate: {metrics.get('Large Move Hit Rate', 0):.2f}%")
    print(f"    - Sharpe Ratio: {metrics.get('Sharpe Ratio', 0):.3f}")
    print(f"    - Win Rate: {metrics.get('Win Rate', 0):.2f}%")

    if 'Avg Uncertainty' in metrics:
        print(f"    - Avg Uncertainty: {metrics['Avg Uncertainty']:.6f}")

    if model_type == 'ensemble':
        print(f"    - Ensemble of {len(model_data['model'].models)} models")

print(f"\n All data cached in {DATA_FILE}")
print(f"   Next run will load from cache automatically!")
print("="*80)